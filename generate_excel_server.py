#!/usr/bin/env python3
"""
DEADLIFT.JP - クライアントExcel生成スクリプト
使い方: python3 generate_excel.py
"""
import sys, json, base64, re, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLACK="0A0A0A"; DARK="1A1A1A"; MID="2A2A2A"; GOLD="C9A84C"; GOLD_L="E8C97A"; WHITE="F0EDE8"; GRAY="888888"

def sf(size=11, bold=False, color=WHITE, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)
def fl(c): return PatternFill("solid", fgColor=c)
def al(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bb(): return Border(bottom=Side(style="thin", color="3A3A3A"))

def sc(ws, row, col, val, font=None, bg=None, aln=None, bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    if font: c.font = font
    if bg:   c.fill = fl(bg)
    if aln:  c.alignment = aln
    if bdr:  c.border = bdr
    return c

def mc(ws, r1, c1, r2, c2, val=None, font=None, bg=None, aln=None):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    if bg:
        for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
            for cell in row:
                try: cell.fill = fl(bg)
                except: pass
    top = ws.cell(row=r1, column=c1)
    if val is not None: top.value = val
    if font: top.font = font
    if aln:  top.alignment = aln
    return top

WEEK_STARTS = [3, 12, 21, 30]

def parse_program(text):
    default = [
        {"label":"DAY 1  スクワット","exercises":[
            {"movement":"コンペティション スクワット","tempo":"","sets":1,"reps":2,"load":"","rpe":"@5-6"},
            {"movement":"コンペティション スクワット","tempo":"","sets":1,"reps":4,"load":"","rpe":"85%"},
            {"movement":"コンペティション スクワット","tempo":"","sets":2,"reps":6,"load":"","rpe":"@7-8"},
            {"movement":"レッグエクステンション","tempo":"","sets":3,"reps":12,"load":"","rpe":"@7-8"},
            {"movement":"ケーブルクランチ","tempo":"","sets":3,"reps":15,"load":"","rpe":"@8"},
        ]},
        {"label":"DAY 2  ベンチプレス","exercises":[
            {"movement":"コンペティション ベンチプレス","tempo":"COMP PAUSE","sets":1,"reps":2,"load":"","rpe":"@5-6"},
            {"movement":"コンペティション ベンチプレス","tempo":"COMP PAUSE","sets":1,"reps":3,"load":"","rpe":"@7"},
            {"movement":"スポト ベンチプレス","tempo":"COMP PAUSE","sets":3,"reps":5,"load":"","rpe":"@7-8"},
            {"movement":"ラーセンプレス","tempo":"","sets":2,"reps":8,"load":"","rpe":"@7-8"},
            {"movement":"ラットプルダウン","tempo":"","sets":3,"reps":10,"load":"","rpe":"@7-8"},
        ]},
        {"label":"DAY 3  デッドリフト","exercises":[
            {"movement":"コンペティション デッドリフト","tempo":"","sets":1,"reps":4,"load":"","rpe":"@3-4"},
            {"movement":"コンペティション デッドリフト","tempo":"","sets":1,"reps":4,"load":"","rpe":"@5-6"},
            {"movement":"コンペティション デッドリフト","tempo":"","sets":2,"reps":6,"load":"","rpe":"@5-6"},
            {"movement":"マシン ハムストリングカール","tempo":"","sets":3,"reps":10,"load":"","rpe":"@7-8"},
            {"movement":"トライセップ プッシュダウン","tempo":"","sets":3,"reps":12,"load":"","rpe":"@7-8"},
        ]},
        {"label":"DAY 4  スクワット/ベンチ","exercises":[
            {"movement":"ポーズ スクワット","tempo":"","sets":3,"reps":3,"load":"","rpe":"@7"},
            {"movement":"コンペティション ベンチプレス","tempo":"COMP PAUSE","sets":1,"reps":2,"load":"","rpe":"@7"},
            {"movement":"クローズグリップ ベンチプレス","tempo":"","sets":3,"reps":8,"load":"","rpe":"@7-8"},
            {"movement":"Tバーロウ","tempo":"","sets":3,"reps":10,"load":"","rpe":"@7-8"},
            {"movement":"バイセップカール","tempo":"","sets":3,"reps":12,"load":"","rpe":"@7-8"},
        ]},
    ]
    if not text or len(text) < 100:
        return default
    days = []
    current_day = None
    for line in text.split("\n"):
        t = line.strip()
        dm = re.match(r"DAY\s*([1-9])", t, re.IGNORECASE)
        if dm:
            if current_day: days.append(current_day)
            current_day = {"label": t.replace("*","").replace("#","").strip(), "exercises": []}
            continue
        if not current_day: continue
        if t.startswith("|") and not re.match(r"^[\|:\-\s]+$", t):
            cells = [c.strip() for c in t.split("|") if c.strip() and not re.match(r"^[-:]+$",c.strip())]
            if len(cells) >= 3:
                mv = cells[0]
                if not mv or re.match(r"^(種目|Movement)", mv): continue
                sr = cells[1]; load_str = cells[2]; rpe_str = cells[3] if len(cells)>3 else ""
                sets=reps=""
                m = re.match(r"(\d+)[×xX](\d+)", sr)
                if m: sets=int(m.group(1)); reps=int(m.group(2))
                load = float(re.sub(r"[^\d.]","",load_str)) if re.search(r"\d",load_str) else ""
                rpe = rpe_str.replace("RPE","").strip()
                if rpe and not rpe.startswith("@") and re.search(r"\d",rpe): rpe="@"+rpe
                current_day["exercises"].append({"movement":mv.upper(),"tempo":"","sets":sets,"reps":reps,"load":load,"rpe":rpe})
    if current_day: days.append(current_day)
    return days if len(days) >= 3 else default

def build_excel(data):
    name = data.get("name","CLIENT")
    age = data.get("age","")
    weight = data.get("weight","")
    competition = data.get("competition","TBD")
    squat = float(data.get("squat",0) or 0)
    bench = float(data.get("bench",0) or 0)
    deadlift = float(data.get("deadlift",0) or 0)
    program_text = data.get("program","")
    days_data = parse_program(program_text)
    today = datetime.date.today()

    wb = Workbook()

    # CLIENT INFO
    ws1 = wb.active; ws1.title = "CLIENT INFO"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = GOLD
    for col,w in {1:3,2:3,3:22,4:12,5:16,6:18,7:4,8:10,9:13,10:13,11:13}.items():
        ws1.column_dimensions[get_column_letter(col)].width = w
    for r in range(1,40): ws1.row_dimensions[r].height = 18

    ws1.row_dimensions[1].height=30
    mc(ws1,1,3,1,11,"DEADLIFT.JP  |  コーチングプログラム",sf(18,True,GOLD),BLACK,al())
    ws1.row_dimensions[5].height=22
    for col,lbl in [(3,"NAME"),(4,"AGE"),(5,"WEIGHT CLASS"),(6,"NEXT COMPETITION"),(8,"SBD MAX"),(9,"SQUAT"),(10,"BENCH"),(11,"DEADLIFT")]:
        sc(ws1,5,col,lbl,sf(10,True,GOLD),MID,al())
    ws1.row_dimensions[6].height=24
    for col,val,fc in [(3,name,WHITE),(4,age,WHITE),(5,weight,WHITE),(6,competition,WHITE),(8,"KG",GRAY),(9,squat,GOLD_L),(10,bench,GOLD_L),(11,deadlift,GOLD_L)]:
        sc(ws1,6,col,val,sf(12,True,fc),DARK,al())
    ws1.row_dimensions[7].height=18
    sc(ws1,7,8,"LB",sf(9,False,GRAY),DARK,al())
    for col,val in [(9,squat),(10,bench),(11,deadlift)]:
        sc(ws1,7,col,round(val*2.204,1) if val else "",sf(10,False,GRAY),DARK,al())
    ws1.row_dimensions[11].height=20
    sc(ws1,11,3,"Schedule",sf(10,True,WHITE),MID,al())
    for i,d in enumerate(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]):
        sc(ws1,11,4+i,d,sf(9,False,WHITE),MID,al())
    sc(ws1,11,11,"Notes",sf(9,False,WHITE),MID,al())
    for row,ex,pri in [(12,"Squat","1"),(13,"Bench Press","2"),(14,"Deadlift","3")]:
        sc(ws1,row,3,ex,sf(10,False,WHITE),DARK,al("left"))
        sc(ws1,row,4,pri,sf(10,True,GOLD),DARK,al())
    ws1.row_dimensions[17].height=20
    mc(ws1,17,3,17,11,"GUIDELINES",sf(10,True,WHITE),MID,al())
    for i,g in enumerate(["クライアントはトップセットの動画を送ること","バックダウンセットは異なる角度から撮影すること","コーチへの返信は24時間以内","怪我が発生した場合は速やかにコーチに報告すること"]):
        r=18+i
        sc(ws1,r,3,"Expectations:" if i==0 else "",sf(9,True,WHITE),DARK,al("left"))
        mc(ws1,r,4,r,11,g,sf(9,False,"AAAAAA"),DARK,al("left",wrap=True))
    sc(ws1,24,3,"RPE",sf(10,True,GOLD),MID,al())
    sc(ws1,24,4,"Definition",sf(10,True,GOLD),MID,al("left"))
    for i,(rpe,desc) in enumerate([(10,"Concentric Failure"),(9.5,"あと少し重くできたが、追加レップは無理"),(9,"あと1レップできる"),(8.5,"あと1レップ、少し重くもできる"),(8,"あと2レップできる"),(7,"あと3レップできる"),(6,"あと4レップできる"),(5,"あと5レップできる")]):
        r=25+i
        sc(ws1,r,3,rpe,sf(10,True,GOLD),DARK,al())
        mc(ws1,r,4,r,11,desc,sf(9,False,WHITE),DARK,al("left"))

    # TRAINING BLOCK
    ws2 = wb.create_sheet("Training Block #1")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = MID
    for col in range(1,40):
        letter = get_column_letter(col)
        offset = next((col-s for s in WEEK_STARTS if 0<=col-s<=7), -1)
        ws2.column_dimensions[letter].width = {0:20,1:9,2:7,3:7,4:10,5:13,6:11}.get(offset, 2)

    ws2.row_dimensions[1].height=30
    mc(ws2,1,3,1,37,"DEADLIFT.JP  |  Training Block #1",sf(16,True,GOLD),BLACK,al())
    ws2.row_dimensions[2].height=20; ws2.row_dimensions[3].height=22
    for s in WEEK_STARTS:
        sc(ws2,2,s,"フェーズ",sf(10,True,GOLD),DARK,al())
        sc(ws2,2,s+2,"残り週数",sf(10,True,GOLD),DARK,al())
        sc(ws2,2,s+4,"ブロック期間",sf(10,True,GOLD),DARK,al())
        sc(ws2,2,s+5,"コーチメモ",sf(10,True,GOLD),DARK,al())
        sc(ws2,3,s+2,8,sf(11,False,WHITE),MID,al())
        sc(ws2,3,s+4,"4週",sf(11,False,WHITE),MID,al())
    ws2.row_dimensions[5].height=28
    for wi,s in enumerate(WEEK_STARTS):
        mc(ws2,5,s,5,s+7,f"WEEK {wi+1}",sf(14,True,GOLD),BLACK,al())
    ws2.row_dimensions[6].height=16
    for wi,s in enumerate(WEEK_STARTS):
        sc(ws2,6,s,(today+datetime.timedelta(weeks=wi)).strftime("%Y/%m/%d"),sf(9,False,GRAY),aln=al())
    ws2.row_dimensions[7].height=20
    for s in WEEK_STARTS:
        for i,h in enumerate(["種目","テンポ","セット","レップ","重量(kg)","強度/RPE","最終RPE"]):
            sc(ws2,7,s+i,h,sf(9,True,GOLD),MID,al())

    cur = 9
    prev_load_rows = {}
    for day in days_data:
        ws2.row_dimensions[cur].height=22
        for s in WEEK_STARTS:
            mc(ws2,cur,s,cur,s+6,day["label"],sf(10,True,BLACK),GOLD,al("left"))
        cur += 1
        for ex in day["exercises"]:
            ws2.row_dimensions[cur].height=18
            pct_m = re.match(r"^(\d+)%$", str(ex.get("rpe","")))
            for s in WEEK_STARTS:
                sc(ws2,cur,s,  ex["movement"],     sf(10,False,WHITE),DARK,al("left"),bb())
                sc(ws2,cur,s+1,ex.get("tempo",""), sf(9,False,WHITE), DARK,al(),bb())
                sc(ws2,cur,s+2,ex.get("sets",""),  sf(10,False,WHITE),DARK,al(),bb())
                sc(ws2,cur,s+3,ex.get("reps",""),  sf(10,False,WHITE),DARK,al(),bb())
                sc(ws2,cur,s+5,ex.get("rpe",""),   sf(10,False,WHITE),DARK,al(),bb())
                sc(ws2,cur,s+6,"",                 sf(10,False,WHITE),DARK,al(),bb())
                load = ex.get("load","")
                if pct_m and (not load or load == ""):
                    pct = float(pct_m.group(1))/100
                    prev_row = prev_load_rows.get(s)
                    if prev_row:
                        lc = get_column_letter(s+4)
                        formula = f"=MROUND({lc}{prev_row}*{pct},2.5)"
                        c = ws2.cell(row=cur, column=s+4, value=formula)
                        c.font=sf(10,True,GOLD_L); c.fill=fl(DARK); c.alignment=al(); c.border=bb()
                    else:
                        sc(ws2,cur,s+4,"",sf(10,False,GOLD_L),DARK,al(),bb())
                else:
                    lv = float(load) if str(load).replace(".","").isdigit() else (load or "")
                    sc(ws2,cur,s+4,lv,sf(10,False,GOLD_L),DARK,al(),bb())
                    if lv != "": prev_load_rows[s] = cur
            cur += 1
        ws2.row_dimensions[cur].height=16
        for s in WEEK_STARTS:
            mc(ws2,cur,s,cur,s+6,"アスリートメモ",sf(9,False,GOLD),BLACK,al("left"))
        cur += 1
        for _ in range(3):
            ws2.row_dimensions[cur].height=18
            for s in WEEK_STARTS:
                mc(ws2,cur,s,cur,s+6,"",bg=MID)
            cur += 1
        cur += 1

    # PR LOG
    ws3 = wb.create_sheet("PR LOG")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = GOLD
    for col,w in {1:3,2:3,3:16,4:14,5:14,6:14,7:14,8:24}.items():
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.row_dimensions[1].height=30
    mc(ws3,1,3,1,8,"DEADLIFT.JP  |  PR LOG",sf(16,True,GOLD),BLACK,al())
    ws3.row_dimensions[3].height=22
    for col,h in [(3,"DATE"),(4,"SQUAT (KG)"),(5,"BENCH (KG)"),(6,"DEADLIFT (KG)"),(7,"TOTAL (KG)"),(8,"NOTES")]:
        sc(ws3,3,col,h,sf(10,True,GOLD),MID,al())
    ws3.row_dimensions[4].height=22
    sc(ws3,4,3,today.strftime("%Y/%m/%d"),sf(10,False,WHITE),DARK,al())
    sc(ws3,4,4,squat,sf(11,True,GOLD_L),DARK,al())
    sc(ws3,4,5,bench,sf(11,True,GOLD_L),DARK,al())
    sc(ws3,4,6,deadlift,sf(11,True,GOLD_L),DARK,al())
    c7=ws3.cell(row=4,column=7,value="=D4+E4+F4")
    c7.font=sf(11,True,GOLD); c7.fill=fl(DARK); c7.alignment=al()
    sc(ws3,4,8,"初期記録",sf(10,False,WHITE),DARK,al("left"))
    for r in range(5,24):
        ws3.row_dimensions[r].height=18
        for col in range(3,9):
            sc(ws3,r,col,"",bg=DARK,aln=al(),bdr=bb())

    return wb

def run_server(port=8765):
    from http.server import BaseHTTPRequestHandler, HTTPServer
    import io

    class Handler(BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            pass  # quiet logging

        def do_OPTIONS(self):
            self.send_response(200)
            self._cors()
            self.end_headers()

        def do_POST(self):
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            try:
                data = json.loads(body)
                wb = build_excel(data)
                buf = io.BytesIO()
                wb.save(buf)
                xlsx_b64 = base64.b64encode(buf.getvalue()).decode()
                payload = json.dumps({"xlsx": xlsx_b64}).encode()
                self.send_response(200)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as e:
                msg = json.dumps({"error": str(e)}).encode()
                self.send_response(500)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(msg)))
                self.end_headers()
                self.wfile.write(msg)

        def _cors(self):
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")

    httpd = HTTPServer(("localhost", port), Handler)
    print(f"✅ Excel生成サーバー起動: http://localhost:{port}")
    print("   coach-dashboard.html の「Excel生成」ボタンが使えます")
    print("   停止: Ctrl+C")
    httpd.serve_forever()


# ============================================================
# クライアントデータ
# ============================================================

CLIENTS = {
    "yasui": {
        "name": "安井 陵",
        "age": "28歳",
        "weight": "74kg級",
        "competition": "第55回富山県パワーリフティング選手権 (2027-03-21)",
        "squat": 217.5,
        "bench": 130,
        "deadlift": 212.5,
        "program": """\
## クライアント分析
74kg級RAW選手・スモウDL・ミディアムSQ。SQは立ち上がり途中（ミドル局面）での失速が弱点で、ボトムでは潰れた経験なし。DLはファーストプルに課題があり、重量が上がると引き始めで失敗するパターン。BPは足を使えている感覚がなく、胸筋量も不足している。両膝半月板損傷のため深屈曲・ブルガリアンランジ禁止、ハックスクワット不可。RPE管理への適応は良好で、スモロフjr1回目で効果あり。現在停滞中のため4週ごとに記録確認して方針調整。

## プログラムの方針
「弱点局面の直接強化＋低疲労蓄積」の4週間ベースブロック。SQのミドル失速にはピンスクワット（ミドル局面スタート）で爆発力を習得させる。DLのファーストプルにはデッドストップを徹底し、毎回スタンス調整を行ってポジション確立を優先する。BPはスポトプレスでボトム強化、ラーセンプレスで足なし感覚を先に習得してから通常ベンチに移行するフローを採用。胸筋量強化のためDBプレスを高レップで積む。膝への負担を避けるためレッグプレスを採用、補助種目は最小限に抑える。

## DAY 1  SQUAT + DEADLIFT
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION SQUAT | 1×2 | 155 | @5-6 |
| COMPETITION SQUAT | 1×3 | 182.5 | @7 |
| COMPETITION SQUAT | 3×3 | 182.5 | @7-8 |
| PIN SQUAT | 3×3 | 165 | @7 |
| COMPETITION DEADLIFT | 1×3 | 162.5 | @5-6 |
| COMPETITION DEADLIFT | 1×3 | 177.5 | @7 |
| COMPETITION DEADLIFT | 2×3 | 177.5 | @7-8 |
| LEG PRESS | 3×12 |  | @7-8 |

## DAY 2  BENCH PRESS
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION BENCH PRESS | 1×2 | 95 | @5-6 |
| COMPETITION BENCH PRESS | 1×3 | 110 | @7 |
| COMPETITION BENCH PRESS | 3×5 | 107.5 | @7-8 |
| SPOTO BENCH PRESS | 3×5 | 97.5 | @7 |
| LARSEN PRESS | 3×8 |  | @7-8 |
| DUMBBELL BENCH PRESS | 3×12 |  | @7-8 |
| LAT PULLDOWN | 3×10 |  | @7-8 |

## DAY 3  SQUAT + BENCH PRESS ACCESSORY
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION SQUAT | 3×5 | 172.5 | @7 |
| PIN SQUAT | 3×3 | 160 | @7 |
| LARSEN PRESS | 3×8 |  | @7 |
| DUMBBELL BENCH PRESS | 3×12 |  | @7-8 |
| CABLE ROW | 3×10 |  | @7-8 |

## DAY 4  DEADLIFT ACCESSORY + BACK
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION DEADLIFT | 3×3 | 170 | @7 |
| ROMANIAN DEADLIFT | 3×8 |  | @7-8 |
| T BAR ROW | 3×10 |  | @7-8 |
| MACHINE HAMSTRING CURL | 3×10 |  | @7-8 |
| TRICEP PUSHDOWN | 3×12 |  | @7-8 |
""",
    },
    "hashiguchi": {
        "name": "橋口廣志",
        "age": "44歳",
        "weight": "93kg級",
        "competition": "全日本パワーリフティング選手権クラシック部門マスターズ (2026-07-18)",
        "squat": 290,
        "bench": 167.5,
        "deadlift": 250,
        "program": """\
## クライアント分析
SQ日本記録保持者・マスターズ全日本チャンピオン（2025年M1 93kg級）というエリートアスリート。SQ・BPはすでに目標に肉薄しており「維持＋微増」で十分だが、DLのみ練習と試合の乖離が大きく明確な課題。スモウDLにおける「体が前に倒れてからSQ的に立つ」動作パターンが試合重量で崩壊し、再現性の低下を招いていると推察。膝付近のミッドプル失速はヒップ先行による上体の過剰前傾が原因の可能性が高い。肉体労働＋週5日という高疲労環境のため、補助種目の過多は厳禁。右肘の問題によりトライセップ系プッシュダウンは全面除外。

## プログラムの方針
「DLファースト・ミニマムアクセサリー」の4週間ブロック。スモウDLの問題局面（膝〜ロックアウト前の失速）に対してポーズDL（膝通過直後で一時停止）を毎セッション組み込み、上体角度の維持とヒップドライブのタイミングを感覚として習得させる。SQは好調な低レップ高セット構成を維持。BPは高レップ低強度で肩・肘への累積疲労を抑制しながら頻度確保。各DAYの補助種目は最大2種目に限定し、翌日のメインリフト品質を保護する。

## DAY 1  SQUAT
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION SQUAT | 1×2 | 220 | @5-6 |
| COMPETITION SQUAT | 1×2 |  | 85% |
| COMPETITION SQUAT | 5×2 | 225 | @7-8 |
| PAUSE COMPETITION SQUAT | 3×3 | 205 | @7 |
| MACHINE HAMSTRING CURL | 3×10 |  | @7-8 |

## DAY 2  BENCH PRESS
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION BENCH PRESS | 3×5 | 130 | @7 |
| LARSEN COMPETITION BENCH PRESS | 3×8 | 115 | @7-8 |
| LAT PULLDOWN | 3×10 |  | @7-8 |

## DAY 3  DEADLIFT
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION DEADLIFT | 1×3 | 180 | @5 |
| COMPETITION DEADLIFT | 1×3 | 195 | @7 |
| COMPETITION DEADLIFT | 1×3 | 205 | @8 |
| PAUSE COMPETITION DEADLIFT | 3×3 | 175 | @7 |
| T BAR ROW | 3×8 |  | @7-8 |
| MACHINE HAMSTRING CURL | 3×10 |  | @7-8 |

## DAY 4  SQUAT / BENCH PRESS
| 種目 | セット×レップ | Load(kg) | RPE |
|---|---|---|---|
| COMPETITION SQUAT | 5×2 | 232.5 | @7 |
| SPOTO COMPETITION BENCH PRESS | 3×5 | 125 | @7 |
| CABLE ROW | 3×10 |  | @7-8 |
| BICEP CURL | 3×12 |  | @7-8 |
""",
    },
}


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--file":
        # ファイルモード: python3 generate_excel_server.py --file client_data.json
        input_file = sys.argv[2] if len(sys.argv) > 2 else "client_data.json"
        try:
            with open(input_file, "r", encoding="utf-8") as f:
                data = json.load(f)
        except FileNotFoundError:
            print(f"Error: {input_file} が見つかりません")
            sys.exit(1)
        wb = build_excel(data)
        name = data.get("name", "CLIENT").replace(" ", "_")
        output = f"{name}_Training_Program.xlsx"
        wb.save(output)
        print(f"✅ 完成: {output}")
    elif len(sys.argv) > 1 and sys.argv[1] == "--client":
        # クライアント直接生成: python3 generate_excel_server.py --client hashiguchi
        key = sys.argv[2] if len(sys.argv) > 2 else ""
        if key not in CLIENTS:
            print(f"Error: クライアント '{key}' が見つかりません")
            print(f"利用可能: {', '.join(CLIENTS.keys())}")
            sys.exit(1)
        data = CLIENTS[key]
        wb = build_excel(data)
        name = data.get("name", "CLIENT").replace(" ", "_")
        output = f"{name}_Training_Program.xlsx"
        wb.save(output)
        print(f"✅ 完成: {output}")
    else:
        # サーバーモード（デフォルト）
        run_server()
